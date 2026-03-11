import logging
import re
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.config import settings

LOGGER = logging.getLogger(__name__)

warnings.filterwarnings('ignore', message='Unknown extension is not supported and will be removed', category=UserWarning)
warnings.filterwarnings('ignore', message='Conditional Formatting extension is not supported and will be removed', category=UserWarning)

_CACHE_DOCS: list[dict[str, Any]] | None = None
_GREETING_TOKENS = {"hi", "hello", "hey", "hii", "hlo", "hola", "yo", "greetings"}
_INTENT_KEYWORDS = {
    "startup_idea": {"startup", "startups", "idea", "ideas", "concept", "concepts", "business", "revenue", "market", "demand", "industry"},
    "analysis_request": {"analysis", "analyses", "report", "reports", "score", "scores", "risk", "status", "founder", "founders", "budget"},
    "call_log": {"call", "calls", "caller", "phone", "conversation", "intent", "topic", "escalated", "summary"},
    "customer_profile": {"customer", "customers", "profile", "profiles", "stage", "interests", "budget"},
    "agent_performance": {"agent", "agents", "handled", "successful", "resolution", "resolutions", "average", "time", "solved"},
    "faq": {"faq", "faqs", "question", "questions", "answer", "answers", "help"},
}
_DEMAND_RANK = {"very high": 4, "high": 3, "medium": 2, "low": 1, "very low": 0}


def _clean_configured_path(value: str | None) -> str:
    cleaned = (value or "").replace("`r", "").replace("`n", "").strip()
    cleaned = cleaned.strip('"').strip("'")
    return cleaned.strip()


def _configured_excel_path() -> Path | None:
    raw = _clean_configured_path(settings.excel_base_path)
    if not raw:
        return None
    return Path(raw).expanduser()


def _candidate_base_paths() -> list[Path]:
    bases: list[Path] = []
    configured = _configured_excel_path()
    if configured and configured.is_dir():
        bases.append(configured)
    home = Path.home()
    project_root = Path(__file__).resolve().parents[3]
    bases.append(project_root)
    bases.append(home / "Desktop")
    bases.append(home / "OneDrive" / "Desktop")

    unique: list[Path] = []
    seen: set[str] = set()
    for path in bases:
        key = str(path)
        if key not in seen:
            unique.append(path)
            seen.add(key)
    return unique


def _find_default_excel_files() -> list[Path]:
    candidates = [
        "Analysis_requests.xlsx",
        "Analysis_requests-1.xlsx",
        "StartUp_ideas_DataBase.xlsx",
        "FAQ's for Agent.xlsx",
        "Call_logger_Database.xlsx",
        "Customer_Profile_Database.xlsx",
        "Call_DataBase_Agents.xlsx",
    ]

    configured = _configured_excel_path()
    if configured and configured.is_file() and configured.exists():
        return [configured]

    found: list[Path] = []
    seen: set[str] = set()
    for base in _candidate_base_paths():
        if not base.exists():
            continue
        for name in candidates:
            path = base / name
            if path.exists():
                key = str(path.resolve())
                if key not in seen:
                    found.append(path)
                    seen.add(key)
    return found


def _normalize_header(value: Any, index: int) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text.lower()).strip("_")
    return text or f"column_{index}"


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _tokenize(text: str) -> set[str]:
    return set(re.findall(r"[a-zA-Z0-9_]+", text.lower()))


def _pick(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key, "").strip()
        if value:
            return value
    return ""


def _money_value(value: str) -> float:
    digits = re.sub(r"[^0-9.]", "", value or "")
    if not digits:
        return 0.0
    try:
        return float(digits)
    except ValueError:
        return 0.0


def _infer_doc_type(path: Path, headers: list[str]) -> str:
    name = path.name.lower()
    header_set = set(headers)
    if "faq" in name or {"question", "answer"}.issubset(header_set):
        return "faq"
    if "idea_name" in header_set or "startup_ideas" in name:
        return "startup_idea"
    if "analysis_id" in header_set:
        return "analysis_request"
    if "caller_name" in header_set or "conversation_summary" in header_set:
        return "call_log"
    if "customer_id" in header_set or "startup_stage" in header_set:
        return "customer_profile"
    if "calls_handled" in header_set or "successful_resolutions" in header_set:
        return "agent_performance"
    return "generic"


def _build_search_text(path: Path, sheet: str, doc_type: str, row: dict[str, str]) -> str:
    fields = [f"{key.replace('_', ' ')}: {value}" for key, value in row.items() if value]
    return f"file: {path.name}; sheet: {sheet}; type: {doc_type}; " + "; ".join(fields)


def _secondary_rank(doc: dict[str, Any]) -> float:
    row = doc["row"]
    doc_type = doc["doc_type"]
    if doc_type == "startup_idea":
        demand = _DEMAND_RANK.get(_pick(row, "market_demand").lower(), 0)
        revenue = _money_value(_pick(row, "potential_revenue"))
        return demand * 1_000_000_000 + revenue
    if doc_type == "analysis_request":
        success = _money_value(_pick(row, "startup_success_score"))
        risk = _money_value(_pick(row, "risk_score"))
        return success * 1_000 - risk
    if doc_type == "agent_performance":
        return _money_value(_pick(row, "successful_resolutions"))
    return 0.0


def _build_docs() -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    for path in _find_default_excel_files():
        try:
            workbook = load_workbook(filename=path, data_only=True, read_only=True)
        except (InvalidFileException, OSError, ValueError):
            continue
        except Exception:
            continue

        try:
            for sheet in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet]
                    rows = worksheet.iter_rows(values_only=True)
                    header_row = next(rows, None)
                except Exception:
                    continue

                if not header_row:
                    continue

                headers = [_normalize_header(value, index) for index, value in enumerate(header_row)]
                doc_type = _infer_doc_type(path, headers)

                for row_number, values in enumerate(rows, start=2):
                    row: dict[str, str] = {}
                    for index, header in enumerate(headers):
                        value = _stringify(values[index] if index < len(values) else "")
                        if value:
                            row[header] = value
                    if not row:
                        continue

                    text = _build_search_text(path, sheet, doc_type, row)
                    docs.append(
                        {
                            "text": text,
                            "tokens": _tokenize(text),
                            "row": row,
                            "row_number": row_number,
                            "source": str(path),
                            "sheet": sheet,
                            "doc_type": doc_type,
                        }
                    )
        finally:
            workbook.close()
    return docs


def _is_greeting(question: str) -> bool:
    tokens = _tokenize(question)
    if not tokens:
        return False
    allowed = _GREETING_TOKENS | {"happy", "ai", "assistant"}
    return len(tokens) <= 4 and tokens.issubset(allowed)


def _detect_intents(question: str) -> set[str]:
    tokens = _tokenize(question)
    intents = {name for name, keywords in _INTENT_KEYWORDS.items() if tokens & keywords}
    lower_question = question.lower().strip()
    faq_prefixes = ("how ", "what ", "where ", "why ", "when ", "who ")
    if {"faq", "faqs", "question", "questions", "answer", "answers"} & tokens or ((lower_question.startswith(faq_prefixes) or lower_question.endswith("?")) and not intents):
        intents.add("faq")
    return intents


def _score_doc(question: str, question_tokens: set[str], intents: set[str], doc: dict[str, Any]) -> float:
    score = float(len(question_tokens & doc["tokens"]))
    row = doc["row"]
    lower_question = question.lower()

    if doc["doc_type"] in intents:
        score += 5.0

    for field in (
        "question",
        "answer",
        "idea_name",
        "industry",
        "startup_idea",
        "founders",
        "caller_name",
        "topic",
        "customer_id",
        "interests",
    ):
        value = row.get(field, "").lower()
        if not value:
            continue
        overlap = len(question_tokens & _tokenize(value))
        score += overlap
        if value and value in lower_question:
            score += 6.0

    if doc["doc_type"] == "faq":
        score += len(question_tokens & _tokenize(_pick(row, "question"))) * 2
    elif doc["doc_type"] == "startup_idea":
        score += _DEMAND_RANK.get(_pick(row, "market_demand").lower(), 0)

    return score


def _retrieve(question: str, docs: list[dict[str, Any]], k: int = 6) -> list[dict[str, Any]]:
    intents = _detect_intents(question)
    question_tokens = _tokenize(question)
    ranked: list[tuple[float, float, dict[str, Any]]] = []

    for doc in docs:
        score = _score_doc(question, question_tokens, intents, doc)
        if score > 0 or (intents and doc["doc_type"] in intents):
            ranked.append((score, _secondary_rank(doc), doc))

    if not ranked:
        filtered = docs
        if intents:
            typed = [doc for doc in docs if doc["doc_type"] in intents]
            if typed:
                filtered = typed
        return filtered[:k]

    ranked.sort(key=lambda item: (item[0], item[1]), reverse=True)
    selected: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int]] = set()
    for _, _, doc in ranked:
        key = (doc["source"], doc["sheet"], doc["row_number"])
        if key in seen:
            continue
        selected.append(doc)
        seen.add(key)
        if len(selected) == k:
            break
    return selected


def _format_doc(doc: dict[str, Any]) -> str:
    row = doc["row"]
    doc_type = doc["doc_type"]

    if doc_type == "faq":
        question = _pick(row, "question")
        answer = _pick(row, "answer")
        category = _pick(row, "category")
        parts = [f"FAQ: {question}", f"Answer: {answer}"]
        if category:
            parts.append(f"Category: {category}")
        return ". ".join(part for part in parts if part)

    if doc_type == "startup_idea":
        return (
            f"Startup idea: {_pick(row, 'idea_name')} | Industry: {_pick(row, 'industry')} | "
            f"Target market: {_pick(row, 'target_market')} | Budget: {_pick(row, 'estimated_budget')} | "
            f"Potential revenue: {_pick(row, 'potential_revenue')} | Market demand: {_pick(row, 'market_demand')} | "
            f"Description: {_pick(row, 'description')}"
        )

    if doc_type == "analysis_request":
        return (
            f"Analysis: {_pick(row, 'analysis_id')} | Founders: {_pick(row, 'founders')} | "
            f"Startup idea: {_pick(row, 'startup_idea')} | Industry: {_pick(row, 'industry')} | "
            f"Budget: {_pick(row, 'budget')} | Analysis type: {_pick(row, 'analysis_type')} | "
            f"Status: {_pick(row, 'status')} | Success score: {_pick(row, 'startup_success_score')} | "
            f"Risk score: {_pick(row, 'risk_score')} | Created date: {_pick(row, 'created_date')}"
        )

    if doc_type == "call_log":
        return (
            f"Call log: {_pick(row, 'call_id')} | Caller: {_pick(row, 'caller_name')} | "
            f"Phone: {_pick(row, 'phone_number')} | Date: {_pick(row, 'call_date')} {_pick(row, 'call_time')} | "
            f"Intent: {_pick(row, 'user_intent')} | Topic: {_pick(row, 'topic')} | "
            f"Summary: {_pick(row, 'conversation_summary')} | Action: {_pick(row, 'action_taken')} | "
            f"Escalated: {_pick(row, 'escalated_to_human')}"
        )

    if doc_type == "customer_profile":
        return (
            f"Customer profile: {_pick(row, 'customer_id')} | Industry: {_pick(row, 'industry')} | "
            f"Startup stage: {_pick(row, 'startup_stage')} | Budget range: {_pick(row, 'budget_range')} | "
            f"Interests: {_pick(row, 'interests')}"
        )

    if doc_type == "agent_performance":
        return (
            f"Agent performance | Date: {_pick(row, 'date')} | Calls handled: {_pick(row, 'calls_handled')} | "
            f"Successful resolutions: {_pick(row, 'successful_resolutions')} | Not solved: {_pick(row, 'not_solved')} | "
            f"Average call time (min): {_pick(row, 'average_call_time_min')}"
        )

    details = ", ".join(f"{key.replace('_', ' ')}: {value}" for key, value in row.items() if value)
    return details or f"Row {doc['row_number']} from {Path(doc['source']).name}"


def _format_sources(docs: list[dict[str, Any]], limit: int = 6) -> list[dict[str, str]]:
    unique: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for doc in docs:
        key = (doc["source"], doc["sheet"])
        if key in seen:
            continue
        unique.append({"source": doc["source"], "sheet": doc["sheet"]})
        seen.add(key)
        if len(unique) == limit:
            break
    return unique


def _startup_idea_response(docs: list[dict[str, Any]]) -> str:
    ideas = [doc for doc in docs if doc["doc_type"] == "startup_idea"]
    if not ideas:
        return ""
    ranked = sorted(ideas, key=lambda doc: _secondary_rank(doc), reverse=True)
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    for doc in ranked:
        key = _pick(doc["row"], "idea_id", "idea_name")
        if key in seen:
            continue
        unique.append(doc)
        seen.add(key)
        if len(unique) == 3:
            break

    lines = ["Here are startup ideas from your Excel database:"]
    for index, doc in enumerate(unique, start=1):
        row = doc["row"]
        lines.append(
            f"{index}. {_pick(row, 'idea_name')} - Industry: {_pick(row, 'industry')}. "
            f"Target market: {_pick(row, 'target_market')}. Budget: {_pick(row, 'estimated_budget')}. "
            f"Potential revenue: {_pick(row, 'potential_revenue')}. Market demand: {_pick(row, 'market_demand')}."
        )
    return "\n".join(lines)


def _faq_response(docs: list[dict[str, Any]]) -> str:
    faq_docs = [doc for doc in docs if doc["doc_type"] == "faq"]
    if not faq_docs:
        return ""
    top = faq_docs[0]["row"]
    answer = _pick(top, "answer")
    question = _pick(top, "question")
    category = _pick(top, "category")
    response = answer or f"Closest FAQ match: {question}"
    extras = []
    if category:
        extras.append(f"Category: {category}")
    if question:
        extras.append(f"Matched FAQ: {question}")
    if extras:
        response = response + "\n" + "\n".join(extras)
    return response


def _analysis_response(docs: list[dict[str, Any]]) -> str:
    analysis_docs = [doc for doc in docs if doc["doc_type"] == "analysis_request"]
    if not analysis_docs:
        return ""
    lines = ["I found these analysis records:"]
    for index, doc in enumerate(analysis_docs[:3], start=1):
        row = doc["row"]
        lines.append(
            f"{index}. {_pick(row, 'analysis_id')} - {_pick(row, 'startup_idea')} by {_pick(row, 'founders')}. "
            f"Industry: {_pick(row, 'industry')}. Budget: {_pick(row, 'budget')}. "
            f"Type: {_pick(row, 'analysis_type')}. Status: {_pick(row, 'status')}. "
            f"Success score: {_pick(row, 'startup_success_score')}. Risk score: {_pick(row, 'risk_score')}."
        )
    return "\n".join(lines)


def _call_log_response(docs: list[dict[str, Any]]) -> str:
    call_docs = [doc for doc in docs if doc["doc_type"] == "call_log"]
    if not call_docs:
        return ""
    lines = ["I found these call log matches:"]
    for index, doc in enumerate(call_docs[:3], start=1):
        row = doc["row"]
        lines.append(
            f"{index}. {_pick(row, 'caller_name')} on {_pick(row, 'call_date')} {_pick(row, 'call_time')}. "
            f"Intent: {_pick(row, 'user_intent')}. Topic: {_pick(row, 'topic')}. "
            f"Summary: {_pick(row, 'conversation_summary')}. Escalated to human: {_pick(row, 'escalated_to_human')}."
        )
    return "\n".join(lines)


def _customer_profile_response(docs: list[dict[str, Any]]) -> str:
    customer_docs = [doc for doc in docs if doc["doc_type"] == "customer_profile"]
    if not customer_docs:
        return ""
    lines = ["I found these customer profiles:"]
    for index, doc in enumerate(customer_docs[:3], start=1):
        row = doc["row"]
        lines.append(
            f"{index}. {_pick(row, 'customer_id')} - Industry: {_pick(row, 'industry')}. "
            f"Stage: {_pick(row, 'startup_stage')}. Budget range: {_pick(row, 'budget_range')}. "
            f"Interests: {_pick(row, 'interests')}."
        )
    return "\n".join(lines)


def _agent_performance_response(docs: list[dict[str, Any]]) -> str:
    agent_docs = [doc for doc in docs if doc["doc_type"] == "agent_performance"]
    if not agent_docs:
        return ""
    lines = ["I found these agent performance records:"]
    for index, doc in enumerate(agent_docs[:3], start=1):
        row = doc["row"]
        lines.append(
            f"{index}. {_pick(row, 'date')} - Calls handled: {_pick(row, 'calls_handled')}. "
            f"Successful resolutions: {_pick(row, 'successful_resolutions')}. Not solved: {_pick(row, 'not_solved')}. "
            f"Average call time: {_pick(row, 'average_call_time_min')} minutes."
        )
    return "\n".join(lines)


def _generic_excel_response(docs: list[dict[str, Any]]) -> str:
    if not docs:
        return "I loaded your Excel files, but I could not find a close match yet. Try asking about startup ideas, FAQs, customer profiles, call logs, or analysis reports."
    lines = ["Here are the closest matches from your Excel sheets:"]
    for index, doc in enumerate(docs[:3], start=1):
        lines.append(f"{index}. {_format_doc(doc)}")
    return "\n".join(lines)


def _fallback_answer(question: str, retrieved: list[dict[str, Any]], ai_status: str) -> str:
    if _is_greeting(question):
        source_names = []
        for source in _format_sources(retrieved, limit=6):
            label = f"{Path(source['source']).name} ({source['sheet']})"
            if label not in source_names:
                source_names.append(label)
        source_text = ", ".join(source_names)
        return (
            "Hi, I am ready to help with your Excel knowledge base. "
            "Try questions like 'give me 3 startup ideas', 'show matching FAQs', 'summarize customer profiles', "
            f"or 'show recent call issues'. Loaded sources: {source_text}."
        )

    intents = _detect_intents(question)
    answer = ""
    if "startup_idea" in intents:
        answer = _startup_idea_response(retrieved)
    elif "analysis_request" in intents:
        answer = _analysis_response(retrieved)
    elif "call_log" in intents:
        answer = _call_log_response(retrieved)
    elif "customer_profile" in intents:
        answer = _customer_profile_response(retrieved)
    elif "agent_performance" in intents:
        answer = _agent_performance_response(retrieved)
    elif "faq" in intents:
        answer = _faq_response(retrieved)

    if not answer:
        answer = _faq_response(retrieved) or _startup_idea_response(retrieved) or _generic_excel_response(retrieved)

    return answer


def _history_to_messages(history: list[dict] | None) -> list[dict[str, str]]:
    messages: list[dict[str, str]] = []
    for item in history or []:
        role = item.get("role")
        content = str(item.get("content", "")).strip()
        if role in {"user", "assistant"} and content:
            messages.append({"role": role, "content": content})
    return messages[-6:]


def _build_context(retrieved: list[dict[str, Any]]) -> str:
    blocks = []
    for index, doc in enumerate(retrieved[:6], start=1):
        blocks.append(
            f"[{index}] Source: {Path(doc['source']).name} | Sheet: {doc['sheet']} | Row: {doc['row_number']}\n{_format_doc(doc)}"
        )
    return "\n\n".join(blocks)


def ask_excel_rag(question: str, history: list[dict] | None = None) -> dict[str, Any]:
    """Excel-only assistant.

    Uses your xlsx files as a knowledge base and answers:
    - FAQs (from the FAQ sheet)
    - startup ideas
    - analysis requests
    - call logs
    - customer profiles
    - agent performance

    No external AI model is used.
    """

    global _CACHE_DOCS
    if _CACHE_DOCS is None:
        _CACHE_DOCS = _build_docs()

    if not _CACHE_DOCS:
        return {
            "answer": (
                "No Excel knowledge files found. Place the listed xlsx files on your Desktop, "
                "or set EXCEL_BASE_PATH in backend/.env to point to the folder or file."
            ),
            "sources": [],
            "mode": "no_data",
        }

    if _is_greeting(question):
        return {
            "answer": _fallback_answer(question, _CACHE_DOCS, ai_status="ready"),
            "sources": _format_sources(_CACHE_DOCS),
            "mode": "greeting",
        }

    intents = _detect_intents(question)

    preferred_types = [
        t
        for t in (
            "faq",
            "startup_idea",
            "analysis_request",
            "call_log",
            "customer_profile",
            "agent_performance",
        )
        if t in intents
    ]

    candidates = _CACHE_DOCS
    if preferred_types:
        filtered = [d for d in _CACHE_DOCS if d.get("doc_type") in preferred_types]
        candidates = filtered or _CACHE_DOCS

    retrieved = _retrieve(question, candidates, k=8)

    # If we have FAQ intent (or only FAQ rows were retrieved), return FAQ answer first.
    if "faq" in intents or any(d.get("doc_type") == "faq" for d in retrieved):
        faq_only = [d for d in retrieved if d.get("doc_type") == "faq"]
        faq_answer = _faq_response(faq_only)
        if faq_answer:
            return {
                "answer": faq_answer,
                "sources": _format_sources(faq_only),
                "mode": "faq",
            }

    answer = _fallback_answer(question, retrieved, ai_status=None)
    return {
        "answer": answer,
        "sources": _format_sources(retrieved),
        "mode": "excel_only",
    }

